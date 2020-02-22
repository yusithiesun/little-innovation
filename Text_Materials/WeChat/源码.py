'''
项目介绍：
1.从xlsx中读取公众号文章url等信息
2.爬取url中的文本以txt形式存入对应文件夹
说明：
因为现阶段只是作为个获取语料的小工具，所以并没有在代码中加入异常处理
'''
import requests
import lxml
import os
from openpyxl import load_workbook
from bs4 import BeautifulSoup

def getText(url):
    '''
    返回网页url中的文案
    '''
    Text=''
    html=requests.get(url)  #获取网页
    bs=BeautifulSoup(html.text,'lxml')  #解析页面
    '''
    #原始方法：直接通过p标签获取文本，会有遗漏
    for p in bs.findAll('p'):
        Text=Text+p.get_text()
    '''
    #通过分析页面结构：更新爬取策略如下
    #放入url
    Text=Text+"文章链接："+url+"\n"
    #爬取标题
    name=bs.find('h2',{'id':'activity-name'})
    Text=Text+"标题：\n"+name.get_text()
    #爬取公众号信息
    info=bs.find('div',{'id':'meta_content'})
    Text=Text+"\n公众号信息：\n"+info.get_text()
    #爬取正文
    Text=Text+"\n正文：\n"
    content=bs.find('div',{'id':'js_content'})
    '''
    #试图分段，目前失败了，有需要再尝试
    if content.find('section') == None:
        for i in content.findAll('p'):
            Text=Text+i.get_text()+"\n"
    else:      
        for i in content.find('section').children:
            Text=Text+i.get_text()+"\n"
    '''
    Text=Text+content.get_text()
    return Text

def writeText(i,text):
    '''
    将文本text写入文件夹中，i为文件名
    '''  
    #print(text)
    fp=open(str(i)+".txt","w",encoding='utf-8')
    fp.write(text)
    fp.close()


path = '***'  #指定目标路径

if __name__=="__main__":
    #设置读取的xlsx路径
    wb=load_workbook('***')
    sheet=wb['Sheet1']

    os.mkdir(path + './***')  #创建文件夹
    path='***'  #切换当前路径
    os.chdir(path)
    #逐行遍历xlsx中的信息，进行爬取
    for i in range(2,sheet.max_row+1):
        if sheet.cell(i,4).value == None:break
        name=sheet.cell(i,1).value+"_"+sheet.cell(i,2).value+"_"+sheet.cell(i,3).value+"_"+str(i-1)
        print(name)
        #url.append(sheet.cell(i,1).value)
        #print(sheet.cell(i,1).value)
        writeText(name,getText(sheet.cell(i,4).value))
