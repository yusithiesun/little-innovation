'''
项目介绍：按SCUT.xlsx中的链接爬取华南理工大学官网讲座预告文本
说明：由于该批网页由jQuery加载，直接爬取会加载不全
（具体表现为包裹正文的<article>标签用requests库get后会缺少结束标签</article>
从而导致抓取的文本会有噪声项干扰）
因而使用selenium库加载js后再爬取
'''
import lxml
import os
#import time    #经试验暂时不需要等待加载，其实偶尔会有来不及加载的情况加载，1136次爬取中出现了2次
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

def writeText(title,text):
    '''
    将文本写入文件夹中
    title为文件名
    text为文件内容
    '''  
    #print(text)
    fp=open(title+".txt","w",encoding='utf-8')
    fp.write(text)
    fp.close()

def getText(url):
    '''
    返回网页中的标题和文案
    '''
    Text=''
    #设置无界面化的chrome浏览器
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    #实例化一个浏览器对象
    browser = webdriver.Chrome(options=chrome_options,executable_path="***/chromedriver")
    #获取加载后的页面
    browser.get(url)
    #解析页面
    bs=BeautifulSoup(browser.page_source,'lxml')
    #放入url
    Text=Text+"文章链接："+url+"\n"
    #爬取标题
    Title=bs.find('title').get_text()
    #爬取正文
    Text=Text+bs.find('article',{'class':'read'}).get_text()
    return [Title,Text]

path = '***'
#不可出现在文件名中的保留字符，将统一替换成'-'
char_set=['/','\\',':','*','?','|','"','>','<']

if __name__=="__main__":
    #设置读取的xlsx路径
    wb=load_workbook('***/SCUT.xlsx')
    sheet=wb['Sheet']

    os.mkdir(path + './SCUT_Text')
    path='C:/Users/Joyce/Desktop/SCUT_Text'
    os.chdir(path)
    
    for i in range(1,sheet.max_row+1):
        if sheet.cell(i,1).value == None:break
        Title,Text=getText(sheet.cell(i,1).value)
        #替换保留字符
        for char in char_set:
            Title=Title.replace(char,'-')
        #print(Title)
        #输出txt
        writeText(str(i)+"_"+Title,Text)
     
