'''
项目介绍：爬取上海大学新闻传播学院所有学术讲座消息，链接存在SHU.xlsx中，文本存在SHU_Text.zip中
说明：
1.起始网页：http://sjc.shu.edu.cn/tg/xsjz.htm
'''
import requests
import lxml
import re
import time 
from bs4 import BeautifulSoup
import os
from openpyxl import Workbook

def writeText(title,text):
    '''
    将文本写入文件夹中
    title为文件名
    '''  
    #print(text)
    fp=open(title+".txt","w",encoding='utf-8')
    fp.write(text)
    fp.close()

def getText(url):
    '''
    返回网页文案
    '''
    Text=''
    Text=Text+url+"\n"
    html=requests.get(url)
    html.encoding='utf-8'
    bs=BeautifulSoup(html.text,'lxml')
    content=bs.find('table',{'id':'dnn_ctr44082_ArtDetail_Table3'})
    return Text+content.get_text()

def getURL(seedURL):
    '''
    返回从起始页开始的所有内链
    '''
    global postPath
    global char_set

    url_list=[]
    # 1.特殊第一页，url格式以及内链格式与其他页不同
    url=seedURL.format(index='')
    html=requests.get(url)
    html.encoding='utf-8'
    bs=BeautifulSoup(html.text,'lxml')
    for link in bs.findAll('a',href=re.compile('^(../info/1008/.*)')):
            Title=link.get_text()
            for char in char_set:   #用空格替换保留字符
                Title=Title.replace(char,' ')
            Address=postPath+link['href'][2:]
            url_list.append([Title,Address])
    time.sleep(1)
    # 2.其他页
    for i in range(4,0,-1): #页数是倒着来的
        url=seedURL.format(index='/'+str(i))
        html=requests.get(url)
        html.encoding='utf-8'
        bs=BeautifulSoup(html.text,'lxml')
        
        for link in bs.findAll('a',href=re.compile('^(../../info/1008/.*)')):
            Title=link.get_text()
            for char in char_set:   #用空格替换保留字符
                Title=Title.replace(char,' ')
            Address=postPath+link['href'][5:]
            #不同页也会出现相同文章，if判断一下，以免重复
            if [Title,Address] not in url_list:url_list.append([Title,Address])
        time.sleep(1)

    return url_list

postPath="http://sjc.shu.edu.cn"  #前缀网址
char_set=['/','\\',':','*','?','|','"','>','<'] #保留字符
if __name__=="__main__":
    path = 'C:/Users/Joyce/Desktop'
    os.mkdir(path + './SHU_Text')
    path='C:/Users/Joyce/Desktop/SHU_Text'
    os.chdir(path)

    seedURL="http://sjc.shu.edu.cn/tg/xsjz{index}.htm"
    url_list=getURL(seedURL)
    try:
        for i in url_list:
            writeText(i[0],getText(i[1])) 
            time.sleep(1)
    finally:
        wb=Workbook()
        sheet=wb.active
        for link,index in zip(url_list,range(1,len(url_list)+1)):
            sheet.cell(index,1).value=link[0]
            sheet.cell(index,2).value=link[1] 
        wb.save("C:/Users/Joyce/Desktop/SHU.xlsx")